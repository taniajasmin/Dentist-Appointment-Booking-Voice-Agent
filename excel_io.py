from openpyxl import Workbook, load_workbook
from pathlib import Path
from typing import Dict, List, Tuple
from datetime import datetime, timedelta

FILE = Path("bookings_dentist.xlsx")   # new filename to keep things clean
HEADERS = ["date", "time", "patient_name", "service", "phone", "status"]
DOCTORS = ["Dr. Jesan Ahmed", "Dr. Hasan Rahman", "Dr. Gregory House"] 

def ensure_workbook_with_doctors():
    """Create workbook with one worksheet per doctor if missing."""
    if FILE.exists():
        return
    wb = Workbook()
    # first sheet
    ws = wb.active
    ws.title = DOCTORS[0]
    ws.append(HEADERS)
    # rest
    for name in DOCTORS[1:]:
        ws2 = wb.create_sheet(title=name)
        ws2.append(HEADERS)
    wb.save(FILE)

def list_doctors() -> List[str]:
    ensure_workbook_with_doctors()
    wb = load_workbook(FILE)
    return wb.sheetnames

def _parse_hhmm(s: str) -> datetime:
    return datetime.strptime(s, "%H:%M")

def _overlaps(t1: str, t2: str) -> bool:
    """Returns True if start times are less than 60 mins apart (1-hour slots)."""
    try:
        a = _parse_hhmm(t1)
        b = _parse_hhmm(t2)
        return abs((a - b).total_seconds()) < 60*60
    except Exception:
        # fall back to strict equality if parse fails
        return t1 == t2

def within_hours(time_str: str) -> bool:
    """Working hours: 14:00–23:59 inclusive."""
    try:
        t = _parse_hhmm(time_str)
        return _parse_hhmm("14:00") <= t <= _parse_hhmm("23:59")
    except Exception:
        return False

def doctor_exists(name: str) -> bool:
    return name in list_doctors()

def slot_available(doctor: str, date_str: str, time_str: str) -> bool:
    """True if no overlapping appointment for that doctor/date/time."""
    ensure_workbook_with_doctors()
    wb = load_workbook(FILE)
    if doctor not in wb.sheetnames:
        return False
    ws = wb[doctor]
    for r in ws.iter_rows(min_row=2, values_only=True):
        d, t, *_ = r
        if str(d) == date_str and _overlaps(str(t), time_str):
            return False
    return True

def append_booking(doctor: str, row: Dict):
    """Write a new row to the doctor's sheet. Creates workbook if needed."""
    ensure_workbook_with_doctors()
    wb = load_workbook(FILE)
    if doctor not in wb.sheetnames:
        # if doctor missing, create sheet with headers (you can also forbid this)
        ws = wb.create_sheet(title=doctor)
        ws.append(HEADERS)
    ws = wb[doctor]
    ws.append([
        row.get("date",""),
        row.get("time",""),
        row.get("patient_name",""),
        row.get("service",""),
        row.get("phone",""),
        row.get("status","confirmed")
    ])
    wb.save(FILE)