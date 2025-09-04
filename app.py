import os
import re
import json
import difflib
from dotenv import load_dotenv

from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, RedirectResponse, HTMLResponse

import dateparser

# Excel I/O helpers (for excel_io.py)
from excel_io import (
    list_doctors,
    doctor_exists,
    slot_available,
    append_booking,
    within_hours,
    ensure_workbook_with_doctors,
)

# Optional: readback for /api/bookings
from openpyxl import load_workbook
from pathlib import Path

load_dotenv()
BRAND = os.getenv("BRAND_NAME", "Demo Dental Clinic")
TZ = os.getenv("TZ", "UTC")

from openai import OpenAI
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

app = FastAPI()

# CORS so the browser frontend can call APIs locally
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="static", html=True), name="static")

@app.get("/", response_class=HTMLResponse)
async def root():
    return RedirectResponse(url="/static/")

# Ensure the Excel file exists once at startup (won't overwrite if present)
@app.on_event("startup")
def _init_file():
    ensure_workbook_with_doctors()

# List doctors (fed to the frontend and NLU)
@app.get("/api/doctors")
def api_doctors():
    return {"doctors": list_doctors()}

def _norm(s: str) -> str:
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"\bdr\.?\b", "", s)       # remove 'dr' / 'dr.'
    s = re.sub(r"[^a-z0-9\s]", " ", s)    # drop punctuation
    return re.sub(r"\s+", " ", s).strip()

def _tokens(s: str):
    return _norm(s).split()

def choose_doctor(user_text: str):
    """
    Return (canonical, None) if confident;
           (None, [choices]) if ambiguous;
           (None, None) if no match.

    Rules:
      - If a token (≥3 chars) uniquely matches the **start** of a doctor's token, choose that doctor.
      - If a token matches multiple doctors (e.g., 'ahmed'), ask to clarify among those.
      - Else try a conservative full-string similarity.
    """
    if not user_text:
        return None, None

    options = list_doctors()

    # Build token -> doctors map
    tok2docs = {}
    for doc in options:
        for t in _tokens(doc):
            tok2docs.setdefault(t, set()).add(doc)

    # User tokens (ignore super short tokens)
    user_toks = [t for t in _tokens(user_text) if len(t) >= 3]

    if user_toks:
        matches = set()
        for ut in user_toks:
            # any doctor token starting with user token
            hits = {doc for tok, docs in tok2docs.items() if tok.startswith(ut) for doc in docs}
            matches |= hits

        if len(matches) == 1:
            return next(iter(matches)), None
        if len(matches) > 1:
            # ambiguous between a few (keep to top 2 alphabetically for short prompt)
            return None, sorted(matches)[:2]

    # fallback: close full-string match (conservative to avoid wrong guesses)
    nu = _norm(user_text)
    scored = [(difflib.SequenceMatcher(None, nu, _norm(doc)).ratio(), doc) for doc in options]
    scored.sort(reverse=True)
    best, name = scored[0]
    if best >= 0.72:
        return name, None

    return None, None

SYSTEM_PROMPT = """
You are a friendly, concise receptionist for a dental clinic.
Your job is to collect exactly 6 fields: doctor, name, phone, service, date, time.
ALWAYS ask for the DOCTOR first. The list of valid doctors will be provided in the user message.
Keep responses short (≤15 words), one question at a time.

Output format (STRICT): respond with pure JSON only (no extra text):
{
  "filled": {"doctor":"", "name":"", "phone":"", "service":"", "date_text":"", "time_text":""},
  "next_question": "string",
  "ready": false
}

Rules:
- If the caller provides multiple fields at once, accept them.
- Phone can include spaces or words (e.g., 'zero'). Convert everything into digits only.
- Accept once at least 8 digits are present.
- Preserve the caller’s wording for date_text and time_text (do not reformat).
- If the spoken doctor name is a partial or close spelling of a listed doctor, set filled.doctor to that listed name.
  If it’s unclear between two doctors, ask “Did you mean Dr. X or Dr. Y?”.
- If the caller says “wait”, “hold on”, or is silent/unclear:
  - set next_question to a brief acknowledgment like “No problem—say ‘I’m ready’ to continue.”
  - keep ready = false and do not lose previously filled fields.
- If the caller says “hello”, “I’m ready”, or “continue”:
  - resume asking for the next missing field (do NOT restart).
- Be natural and warm. If asked, you are the clinic’s automated assistant.
- When all 6 fields are present, set ready = true and ask a short confirmation question.

Behavior:
- Always ask for exactly one missing field per turn (except a final confirmation).
- Never include anything except the JSON object described above.
"""

# Utils
def clean_name(s: str) -> str:
    if not s: return ""
    s = re.sub(r"[^a-zA-Z\s\-]", "", s)  # keep only letters, spaces, hyphens, and capitalize nicely
    return s.strip().title()

def clean_phone(s: str) -> str:
    digits = re.sub(r"\D", "", s or "")
    return digits if len(digits) >= 8 else ""

def normalize(date_text: str, time_text: str, tz_name: str):
    """Return (YYYY-MM-DD, HH:MM) or (None, None) if unparseable."""
    try:
        dt = dateparser.parse(
            f"{date_text} {time_text}",
            settings={"TIMEZONE": tz_name, "RETURN_AS_TIMEZONE_AWARE": False},
        )
        if not dt:
            return None, None
        return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M")
    except Exception:
        return None, None

def next_missing_question(filled: dict) -> str:
    order = [
        ("doctor", "Which doctor would you like to book with?"),
        ("name", "What is your name?"),
        ("phone", "What is your phone number? Please say at least 8 digits."),
        ("service", "What service do you need?"),
        ("date_text", "What date would you like to schedule your appointment?"),
        ("time_text", "What time would you like to schedule your appointment?"),
    ]
    for k, q in order:
        v = (filled.get(k) or "").strip()
        if not v:
            return q
    return "Great—shall I proceed to book this appointment?"

# NLU
@app.post("/nlu")
async def nlu(req: Request):
    """Takes current filled fields + latest user text, returns JSON (fields/next_question/ready)."""
    body = await req.json()
    filled = body.get("filled", {}) or {}
    user_text = body.get("user_text", "") or ""
    doctors = list_doctors()

    # sanitize existing phone
    filled["phone"] = clean_phone(filled.get("phone", ""))

    prompt_user = (
        f"Valid doctors list: {doctors}\n"
        f"Current filled fields: {json.dumps(filled)}\n"
        f'User said: "{user_text}"\n'
        f"Return JSON as specified."
    )

    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0.1,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": prompt_user},
            ],
        )
        raw = (resp.choices[0].message.content or "").strip()
        data = json.loads(raw)
    except Exception:
        data = {
            "filled": filled,
            "next_question": "Sorry, could you repeat that?",
            "ready": False,
        }

    # sanitize patient name again
    if "filled" in data and data["filled"].get("name"):
        data["filled"]["name"] = clean_name(data["filled"]["name"])

    # sanitize phone again
    if "filled" in data and data["filled"].get("phone"):
        data["filled"]["phone"] = clean_phone(data["filled"]["phone"])

    # Fuzzy map spoken doctor to canonical (or ask to clarify if ambiguous)
    doc_in = (data.get("filled", {}) or {}).get("doctor", "")
    if doc_in:
        canon, ambiguous = choose_doctor(doc_in)
        if canon:
            data["filled"]["doctor"] = canon
        elif ambiguous:
            data["filled"]["doctor"] = ""
            data["ready"] = False
            data["next_question"] = f"Did you mean {ambiguous[0]} or {ambiguous[1]}?"
            return JSONResponse(data)  # wait for clarification

    # Final guard: reprompt if still invalid
    doc = (data.get("filled", {}) or {}).get("doctor", "")
    if doc and not doctor_exists(doc):
        data["filled"]["doctor"] = ""
        data["ready"] = False
        data["next_question"] = (
            f"Sorry, that doctor is not in our clinic. "
            f"Available doctors: {', '.join(doctors)}. Which doctor would you like?"
        )
        return JSONResponse(data)

    # Phone must be >=8 digits; otherwise clear it and reprompt
    if (data.get("filled") or {}).get("phone"):
        if len(data["filled"]["phone"]) < 8:
            data["filled"]["phone"] = ""
            data["ready"] = False
            data["next_question"] = "Please say your phone number with at least 8 digits."
            return JSONResponse(data)

    # Choose the next missing field deterministically
    data["next_question"] = next_missing_question(data.get("filled", {}))
    all_ok = all((data["filled"].get(k) or "").strip() for k in ["doctor","name","phone","service","date_text","time_text"])
    data["ready"] = bool(all_ok)

    return JSONResponse(data)


# Booking checks & writing
@app.post("/check")
async def check(req: Request):
    body = await req.json()
    doctor_raw = (body.get("doctor") or "").strip()
    date_text = (body.get("date_text") or "").strip()
    time_text = (body.get("time_text") or "").strip()

    canon, ambiguous = choose_doctor(doctor_raw) if doctor_raw else (None, None)
    if ambiguous:
        return {
            "ok": False,
            "reason": "ambiguous-doctor",
            "message": f"Did you mean {ambiguous[0]} or {ambiguous[1]}?",
        }
    doctor = canon or doctor_raw

    if not doctor_exists(doctor):
        return {
            "ok": False,
            "reason": "unknown-doctor",
            "message": f"Doctor not found. Available: {', '.join(list_doctors())}",
        }

    date_str, time_str = normalize(date_text, time_text, TZ)
    if not date_str or not time_str:
        return {
            "ok": False,
            "reason": "bad-datetime",
            "message": "Sorry, I couldn't understand that date and time.",
        }

    if not within_hours(time_str):
        return {
            "ok": False,
            "reason": "outside-hours",
            "message": "Our doctors are available 14:00 to 23:59. Please choose a time in that range.",
        }

    if not slot_available(doctor, date_str, time_str):
        return {
            "ok": False,
            "reason": "overlap",
            "message": "That time is already booked. Please choose another time or another date.",
        }

    return {"ok": True, "date": date_str, "time": time_str}

@app.post("/book")
async def book(req: Request):
    body = await req.json()
    doctor_raw = (body.get("doctor") or "").strip()
    name = (body.get("name") or "").strip()
    phone = clean_phone(body.get("phone", ""))
    service = (body.get("service") or "").strip()
    date_text = (body.get("date_text") or "").strip()
    time_text = (body.get("time_text") or "").strip()

    canon, ambiguous = choose_doctor(doctor_raw) if doctor_raw else (None, None)
    if ambiguous:
        return JSONResponse({"ok": False, "message": f"Did you mean {ambiguous[0]} or {ambiguous[1]}?"}, status_code=400)
    doctor = canon or doctor_raw

    if not doctor_exists(doctor):
        return JSONResponse(
            {"ok": False, "message": f"Doctor not found. Available: {', '.join(list_doctors())}"},
            status_code=400,
        )

    date_str, time_str = normalize(date_text, time_text, TZ)
    if not date_str or not time_str:
        return JSONResponse({"ok": False, "message": "Invalid date/time."}, status_code=400)

    if not within_hours(time_str):
        return JSONResponse(
            {"ok": False, "message": "Doctors are available 14:00–23:59 only."},
            status_code=400,
        )

    if not slot_available(doctor, date_str, time_str):
        return JSONResponse(
            {"ok": False, "message": "That time is already booked. Please choose another."},
            status_code=409,
        )

    # Append to doctor's sheet
    append_booking(
        doctor,
        {
            "date": date_str,
            "time": time_str,
            "patient_name": name,
            "service": service,
            "phone": phone,
            "status": "confirmed",
        },
    )
    return {"ok": True, "message": f"Booked with {doctor} on {date_str} at {time_str}."}

# View bookings in browser
@app.get("/api/bookings")
def api_bookings(doctor: str = "", date: str = ""):
    """Quick way to inspect rows without opening Excel."""
    from excel_io import FILE  # use the same path as excel_io
    f = Path(FILE)
    if not f.exists():
        return {"rows": []}

    wb = load_workbook(f)
    sheets = [doctor] if doctor and doctor in wb.sheetnames else wb.sheetnames

    all_rows = []
    for sh in sheets:
        ws = wb[sh]
        headers = [c.value for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, r))
            row["doctor"] = sh
            if date and str(row.get("date")) != date:
                continue
            all_rows.append(row)
    return {"rows": all_rows}
